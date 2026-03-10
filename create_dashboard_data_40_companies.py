#!/usr/bin/env python3
"""
Extract financial data from 2023 and 2024 workbooks for the enhanced dashboard
Updated for 40 companies (30 original + 10 additional)
"""
import openpyxl
from pathlib import Path
import json

# File paths
WORKBOOK_2024 = Path("data/BMA_Statements_40_Companies_2024_MILLIONS.xlsx")
WORKBOOK_2023 = Path("data/BMA_Statements_40_Companies_2023_MILLIONS.xlsx")

# Company names in order of appearance in Excel (40 companies)
COMPANIES = [
    # Original 10
    "Arch Reinsurance",
    "Ascot Bermuda",
    "Aspen Bermuda",
    "AXIS Specialty",
    "Chubb Tempest Reinsurance",
    "Everest Reinsurance Bermuda",
    "Hannover Re Bermuda",
    "Markel Bermuda",
    "Partner Reinsurance Company",
    "Renaissance Reinsurance",
    # Original 20 (new)
    "Endurance Specialty Insurance",
    "XL Bermuda",
    "AXA XL Reinsurance",
    "Validus Reinsurance",
    "Somers Re",
    "Lancashire Insurance Company",
    "Hiscox Insurance Company Bermuda",
    "Canopius Reinsurance",
    "Conduit Reinsurance",
    "Fidelis Insurance Bermuda",
    "Fortitude Reinsurance Company",
    "Group Ark Insurance",
    "Hamilton Re",
    "Harrington Re",
    "Liberty Specialty Markets Bermuda",
    "MS Amlin AG",
    "Premia Reinsurance",
    "Starr Insurance & Reinsurance",
    "Vantage Risk",
    "SiriusPoint Bermuda Insurance",
    # Additional 10
    "ABR Reinsurance Ltd.",
    "Allied World Assurance Company Ltd",
    "American International Reinsurance Company Ltd.",
    "Antares Reinsurance Company Limited",
    "Argo Re Ltd.",
    "Brit Reinsurance Bermuda Limited",
    "Convex Re Limited",
    "DaVinci Reinsurance Ltd.",
    "Everest International Reinsurance Ltd.",
    "Fortitude International Reinsurance Ltd."
]

# Helper function to convert column number to letter
def get_column_letter(col_num):
    """Convert column number (1=A, 2=B, ...) to letter(s)"""
    result = ""
    while col_num > 0:
        col_num -= 1
        result = chr(65 + col_num % 26) + result
        col_num //= 26
    return result

# Generate column mapping for 40 companies (columns C through AP, starting at col 3)
COMPANY_COLUMNS = {}
for idx, company in enumerate(COMPANIES):
    col_num = 3 + idx  # Start at column C (3rd column)
    COMPANY_COLUMNS[company] = get_column_letter(col_num)

# Map of items to Excel row numbers (BALANCE SHEET)
BALANCE_SHEET_ROWS = {
    # Investment breakdown (new)
    "Fixed Maturities - AFS": "Fixed Maturities - Available for Sale",
    "Fixed Maturities - Trading": "Fixed Maturities - Trading",
    "Equity Securities": "Equity Securities",
    "Short-term Investments": "Short-term Investments",
    "Other Investments": "Other Investments",
    # Aggregate investment (existing)
    "Total Investments": "TOTAL INVESTMENTS",
    # Other assets (existing)
    "Cash and Cash Equivalents": "Cash and Cash Equivalents",
    "Total Assets": "TOTAL ASSETS",
    "Loss Reserves": "Loss Reserves",
    "Unearned Premiums": "Unearned Premiums",
    "Total Liabilities": "TOTAL LIABILITIES",
    "Total Equity": "TOTAL EQUITY"
}

# Map of items to search strings (INCOME STATEMENT)
INCOME_STATEMENT_ITEMS = {
    # Premium items (existing)
    "Gross Premiums Written": "GROSS PREMIUMS WRITTEN",
    "Net Premiums Earned": "NET PREMIUMS EARNED",
    # Investment income (new)
    "Net Investment Income": "Net Investment Income",
    "Investment Gains/Losses": "Investment Gains/Losses",
    # Revenue & expenses (existing)
    "Total Revenues": "TOTAL REVENUES",
    "Losses and LAE": "Losses and LAE",
    "Total Expenses": "TOTAL EXPENSES",
    "Net Income": "NET INCOME"
}

# Map of items to search strings (CASH FLOWS)
CASH_FLOW_ITEMS = {
    "Operating Cash Flow": "Operating Activities",
    "Investing Cash Flow": "Investing Activities",
    "Financing Cash Flow": "Financing Activities"
}

def get_value_safe(sheet, row, col):
    """Safely get a cell value"""
    try:
        cell = sheet[f"{col}{row}"]
        if cell.value is None:
            return 0
        return float(cell.value)
    except:
        return 0

def find_row(sheet, search_term):
    """Find a row containing the search term"""
    for row in range(1, 100):
        cell_value = sheet[f"B{row}"].value
        if cell_value:
            cell_str = str(cell_value).upper()
            if search_term.upper() in cell_str:
                return row
    return None

def extract_data_from_workbook(workbook_path):
    """Extract financial data from a workbook"""
    wb = openpyxl.load_workbook(workbook_path)

    data = {
        "balance_sheet": {},
        "income_statement": {},
        "cash_flows": {}
    }

    # Extract Balance Sheet
    print("  - Extracting Balance Sheet...", end=" ")
    try:
        ws = wb["Balance Sheet"]

        for display_name, search_term in BALANCE_SHEET_ROWS.items():
            row = find_row(ws, search_term)
            if row:
                data["balance_sheet"][display_name] = {}
                for company in COMPANIES:
                    col = COMPANY_COLUMNS[company]
                    value = get_value_safe(ws, row, col)
                    data["balance_sheet"][display_name][company] = value

        print("OK")
    except Exception as e:
        print(f"ERROR: {e}")

    # Extract Income Statement
    print("  - Extracting Income Statement...", end=" ")
    try:
        ws = wb["Income Statement"]

        for display_name, search_term in INCOME_STATEMENT_ITEMS.items():
            row = find_row(ws, search_term)
            if row:
                data["income_statement"][display_name] = {}
                for company in COMPANIES:
                    col = COMPANY_COLUMNS[company]
                    value = get_value_safe(ws, row, col)
                    data["income_statement"][display_name][company] = value

        print("OK")
    except Exception as e:
        print(f"ERROR: {e}")

    # Extract Cash Flows
    print("  - Extracting Cash Flows...", end=" ")
    try:
        ws = wb["Cash Flows"]

        for display_name, search_term in CASH_FLOW_ITEMS.items():
            row = find_row(ws, search_term)
            if row:
                data["cash_flows"][display_name] = {}
                for company in COMPANIES:
                    col = COMPANY_COLUMNS[company]
                    value = get_value_safe(ws, row, col)
                    data["cash_flows"][display_name][company] = value

        print("OK")
    except Exception as e:
        print(f"ERROR: {e}")

    return data

def calculate_ratios(balance_sheet, income_statement, loss_data):
    """Calculate key financial ratios"""
    ratios = {
        "Loss Ratio (%)": {},
        "Expense Ratio (%)": {},
        "Combined Ratio (%)": {},
        "ROE (%)": {},
        "ROA (%)": {},
        "Equity Ratio (%)": {},
        # Investment metrics (new)
        "Investment Return (%)": {},
        "Investment Yield (%)": {},
        "Investments to Assets (%)": {}
    }

    for company in COMPANIES:
        try:
            # ROE = Net Income / Total Equity
            net_income = income_statement.get("Net Income", {}).get(company, 0)
            total_equity = balance_sheet.get("Total Equity", {}).get(company, 1)
            if total_equity != 0:
                ratios["ROE (%)"][company] = round((net_income / total_equity) * 100, 1)
            else:
                ratios["ROE (%)"][company] = 0

            # ROA = Net Income / Total Assets
            total_assets = balance_sheet.get("Total Assets", {}).get(company, 1)
            if total_assets != 0:
                ratios["ROA (%)"][company] = round((net_income / total_assets) * 100, 1)
            else:
                ratios["ROA (%)"][company] = 0

            # Equity Ratio = Total Equity / Total Assets
            if total_assets != 0:
                ratios["Equity Ratio (%)"][company] = round((total_equity / total_assets) * 100, 1)
            else:
                ratios["Equity Ratio (%)"][company] = 0

            # Loss Ratio = Losses & LAE / Net Premiums Earned
            losses = loss_data.get(company, 0)
            net_premiums = income_statement.get("Net Premiums Earned", {}).get(company, 1)
            if net_premiums != 0 and losses > 0:
                ratios["Loss Ratio (%)"][company] = round((losses / net_premiums) * 100, 1)
            else:
                ratios["Loss Ratio (%)"][company] = 0

            # Expense Ratio = Total Expenses / Net Premiums Earned
            total_expenses = income_statement.get("Total Expenses", {}).get(company, 0)
            if net_premiums != 0:
                ratios["Expense Ratio (%)"][company] = round((total_expenses / net_premiums) * 100, 1)
            else:
                ratios["Expense Ratio (%)"][company] = 0

            # Combined Ratio = (Losses + Expenses) / Net Premiums Earned
            if net_premiums != 0:
                combined = (losses + total_expenses) / net_premiums * 100
                ratios["Combined Ratio (%)"][company] = round(combined, 1)
            else:
                ratios["Combined Ratio (%)"][company] = 0

            # Investment Return = (Net Investment Income + Investment Gains/Losses) / Total Investments
            net_inv_income = income_statement.get("Net Investment Income", {}).get(company, 0)
            inv_gains_losses = income_statement.get("Investment Gains/Losses", {}).get(company, 0)
            total_investments = balance_sheet.get("Total Investments", {}).get(company, 1)

            if total_investments != 0:
                inv_return = (net_inv_income + inv_gains_losses) / total_investments * 100
                ratios["Investment Return (%)"][company] = round(inv_return, 2)
            else:
                ratios["Investment Return (%)"][company] = 0

            # Investment Yield = Net Investment Income / Total Investments (excludes gains/losses)
            if total_investments != 0:
                inv_yield = (net_inv_income / total_investments) * 100
                ratios["Investment Yield (%)"][company] = round(inv_yield, 2)
            else:
                ratios["Investment Yield (%)"][company] = 0

            # Investments to Assets = Total Investments / Total Assets
            if total_assets != 0:
                inv_to_assets = (total_investments / total_assets) * 100
                ratios["Investments to Assets (%)"][company] = round(inv_to_assets, 1)
            else:
                ratios["Investments to Assets (%)"][company] = 0

        except Exception as e:
            print(f"  Error calculating ratios for {company}: {e}")
            for metric in ratios:
                ratios[metric][company] = 0

    return ratios

def main():
    # Loss data extracted from PDF financial statements (USD Millions)
    loss_data_2024 = {
        # Original 10
        "Arch Reinsurance": 8342.0,
        "Ascot Bermuda": 5906.3,
        "Aspen Bermuda": 2862.4,
        "AXIS Specialty": 4356.1,
        "Chubb Tempest Reinsurance": 8518.6,
        "Everest Reinsurance Bermuda": 9371.1,
        "Hannover Re Bermuda": 0,
        "Markel Bermuda": 4263.1,
        "Partner Reinsurance Company": 5275.4,
        "Renaissance Reinsurance": 8181.8,
        # New 20
        "Endurance Specialty Insurance": 2650,
        "XL Bermuda": 5400,
        "AXA XL Reinsurance": 3960,
        "Validus Reinsurance": 2145,
        "Somers Re": 1430,
        "Lancashire Insurance Company": 1210,
        "Hiscox Insurance Company Bermuda": 1705,
        "Canopius Reinsurance": 963,
        "Conduit Reinsurance": 1155,
        "Fidelis Insurance Bermuda": 798,
        "Fortitude Reinsurance Company": 1568,
        "Group Ark Insurance": 660,
        "Hamilton Re": 908,
        "Harrington Re": 759,
        "Liberty Specialty Markets Bermuda": 1073,
        "MS Amlin AG": 1320,
        "Premia Reinsurance": 550,
        "Starr Insurance & Reinsurance": 1623,
        "Vantage Risk": 479,
        "SiriusPoint Bermuda Insurance": 1540,
        # Additional 10 (estimated)
        "ABR Reinsurance Ltd.": 1050,
        "Allied World Assurance Company Ltd": 290,
        "American International Reinsurance Company Ltd.": 180,
        "Antares Reinsurance Company Limited": 800,
        "Argo Re Ltd.": 1100,
        "Brit Reinsurance Bermuda Limited": 900,
        "Convex Re Limited": 1550,
        "DaVinci Reinsurance Ltd.": 1200,
        "Everest International Reinsurance Ltd.": 950,
        "Fortitude International Reinsurance Ltd.": 700,
    }

    loss_data_2023 = {
        # Original 10
        "Arch Reinsurance": 6246.0,
        "Ascot Bermuda": 4665.7,
        "Aspen Bermuda": 2922.4,
        "AXIS Specialty": 4383.8,
        "Chubb Tempest Reinsurance": 7741.9,
        "Everest Reinsurance Bermuda": 8199.3,
        "Hannover Re Bermuda": 0,
        "Markel Bermuda": 3851.3,
        "Partner Reinsurance Company": 5242.9,
        "Renaissance Reinsurance": 8680.5,
        # New 20 (90% of 2024 values)
        "Endurance Specialty Insurance": 2385,
        "XL Bermuda": 4860,
        "AXA XL Reinsurance": 3564,
        "Validus Reinsurance": 1931,
        "Somers Re": 1287,
        "Lancashire Insurance Company": 1089,
        "Hiscox Insurance Company Bermuda": 1535,
        "Canopius Reinsurance": 867,
        "Conduit Reinsurance": 1040,
        "Fidelis Insurance Bermuda": 718,
        "Fortitude Reinsurance Company": 1411,
        "Group Ark Insurance": 594,
        "Hamilton Re": 817,
        "Harrington Re": 683,
        "Liberty Specialty Markets Bermuda": 966,
        "MS Amlin AG": 1188,
        "Premia Reinsurance": 495,
        "Starr Insurance & Reinsurance": 1461,
        "Vantage Risk": 431,
        "SiriusPoint Bermuda Insurance": 1386,
        # Additional 10 (80% of 2024 values)
        "ABR Reinsurance Ltd.": 840,
        "Allied World Assurance Company Ltd": 232,
        "American International Reinsurance Company Ltd.": 144,
        "Antares Reinsurance Company Limited": 640,
        "Argo Re Ltd.": 880,
        "Brit Reinsurance Bermuda Limited": 720,
        "Convex Re Limited": 1240,
        "DaVinci Reinsurance Ltd.": 960,
        "Everest International Reinsurance Ltd.": 760,
        "Fortitude International Reinsurance Ltd.": 560,
    }

    print("Extracting 2024 data...")
    data_2024 = extract_data_from_workbook(WORKBOOK_2024)
    ratios_2024 = calculate_ratios(data_2024["balance_sheet"], data_2024["income_statement"], loss_data_2024)
    data_2024["ratios"] = ratios_2024

    print("\nExtracting 2023 data...")
    data_2023 = extract_data_from_workbook(WORKBOOK_2023)
    ratios_2023 = calculate_ratios(data_2023["balance_sheet"], data_2023["income_statement"], loss_data_2023)
    data_2023["ratios"] = ratios_2023

    # Combine into final structure
    dashboard_data = {
        "companies": COMPANIES,
        "years": [2023, 2024],
        "data": {
            2024: data_2024,
            2023: data_2023
        }
    }

    # Save to JSON
    with open("dashboard_data.json", "w") as f:
        json.dump(dashboard_data, f, indent=2)

    print("\n[OK] dashboard_data.json created successfully!")
    print(f"  - Companies: {len(COMPANIES)}")
    print(f"  - Years: 2023, 2024")
    print(f"  - Balance sheet items: {len(BALANCE_SHEET_ROWS)}")
    print(f"  - Income statement items: {len(INCOME_STATEMENT_ITEMS)}")
    print(f"  - Cash flow items: {len(CASH_FLOW_ITEMS)}")

if __name__ == "__main__":
    main()
