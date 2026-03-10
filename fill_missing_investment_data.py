#!/usr/bin/env python3
"""
Fill in missing investment data in Excel files
Uses industry-standard yields based on actual company data
"""

import json
import openpyxl
from openpyxl.utils import get_column_letter

print("=" * 90)
print("UPDATING EXCEL FILES WITH ESTIMATED INVESTMENT INCOME")
print("=" * 90)

# Load current dashboard data to get company structure
with open('dashboard_data.json', 'r') as f:
    data = json.load(f)

companies = data['companies']
inv_income_2024 = data['data']['2024']['income_statement'].get('Net Investment Income', {})
total_investments_2024 = data['data']['2024']['balance_sheet'].get('Total Investments', {})
inv_income_2023 = data['data']['2023']['income_statement'].get('Net Investment Income', {})
total_investments_2023 = data['data']['2023']['balance_sheet'].get('Total Investments', {})

# Calculate statistics for yield estimation
companies_with_2024_income = {c: inv_income_2024.get(c, 0) for c in companies
                              if inv_income_2024.get(c, 0) > 0}
avg_yield_2024 = (sum((inv_income_2024.get(c, 0) / total_investments_2024.get(c, 1) * 100)
                       for c in companies_with_2024_income
                       if total_investments_2024.get(c, 0) > 0) / len(companies_with_2024_income)
                  if companies_with_2024_income else 3.0)

companies_with_2023_income = {c: inv_income_2023.get(c, 0) for c in companies
                              if inv_income_2023.get(c, 0) > 0}
avg_yield_2023 = (sum((inv_income_2023.get(c, 0) / total_investments_2023.get(c, 1) * 100)
                       for c in companies_with_2023_income
                       if total_investments_2023.get(c, 0) > 0) / len(companies_with_2023_income)
                  if companies_with_2023_income else 3.0)

print(f"\n2024 Average Yield (from actual data): {avg_yield_2024:.2f}%")
print(f"2023 Average Yield (from actual data): {avg_yield_2023:.2f}%")

# Process both Excel files
for filename in ['data/BMA_Statements_30_Companies_2024_MILLIONS.xlsx',
                 'data/BMA_Statements_30_Companies_2023_MILLIONS.xlsx']:

    print(f"\n\nUpdating: {filename}")
    print("-" * 90)

    wb = openpyxl.load_workbook(filename)
    ws = wb['Income Statement']

    # Row 8 = Net Investment Income, Row 9 = Investment Gains/Losses
    year = 2024 if '2024' in filename else 2023
    avg_yield = avg_yield_2024 if year == 2024 else avg_yield_2023
    actual_inv_income = inv_income_2024 if year == 2024 else inv_income_2023
    actual_investments = total_investments_2024 if year == 2024 else total_investments_2023

    # Get company names from header row
    company_columns = {}
    for col in range(3, 33):  # Columns C through AF
        company_name = ws.cell(2, col).value
        if company_name:
            company_columns[company_name] = col

    print(f"\nYear: {year} | Estimated Yield: {avg_yield:.2f}%\n")

    updates_made = 0
    for company_name, col in sorted(company_columns.items()):
        current_income = ws.cell(8, col).value
        investments = actual_investments.get(company_name, 0)

        if current_income is None or current_income == 0:
            # Calculate estimated income
            estimated_income = round(investments * avg_yield / 100, 1)

            if estimated_income > 0:
                ws.cell(8, col).value = estimated_income
                print(f"  {company_name:<35} -> Estimated Income: ${estimated_income:>8,.0f}M")
                updates_made += 1

        # Investment Gains/Losses - set to 0 for all (conservative assumption)
        current_gains = ws.cell(9, col).value
        if current_gains is None or current_gains == 0:
            ws.cell(9, col).value = 0

    print(f"\nTotal updates made to Income Statement: {updates_made}")

    # Save updated workbook
    wb.save(filename)
    print(f"Saved: {filename}")

print("\n" + "=" * 90)
print("NEXT STEPS:")
print("=" * 90)
print("""
1. Run: python create_dashboard_data_multi_year.py
   (to regenerate dashboard_data.json with new estimates)

2. Run: python validate_financial_data.py
   (to validate the updated data)

3. Update DATA_QUALITY_REPORT.md to document:
   - Estimated income methodology
   - Yield estimates used
   - Companies with estimated vs. actual data

4. Run: python test_phase7.py
   (to verify all tests still pass)

5. Commit and push changes to GitHub
""")

EOF
