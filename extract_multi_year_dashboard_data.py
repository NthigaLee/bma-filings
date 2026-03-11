#!/usr/bin/env python3
"""
Flexible Multi-Year Financial Data Extraction
BMA Filings Financial Dashboard

Supports extracting data for any year range and company set from Excel workbooks.
Replaces hardcoded year/company logic with dynamic configuration.
"""

import openpyxl
from pathlib import Path
import json
from datetime import datetime

# ============================================================================
# CONFIGURATION - EDIT HERE TO SUPPORT DIFFERENT YEARS/COMPANIES
# ============================================================================

EXTRACTION_CONFIG = {
    # Years to extract (supports any range)
    # Change this to [2021, 2022, 2023, 2024] for historical expansion
    'years': [2023, 2024],

    # Workbook sources for each year
    'workbook_sources': {
        2021: 'data/by_year/BMA_Class4_2021.xlsx',
        2022: 'data/by_year/BMA_Class4_2022.xlsx',
        2023: 'data/BMA_Statements_40_Companies_2023_MILLIONS.xlsx',
        2024: 'data/BMA_Statements_40_Companies_2024_MILLIONS.xlsx',
    },

    # Target companies (40 Bermuda insurance/reinsurance companies)
    'target_companies': [
        # Original 10
        "Arch Reinsurance",
        "Ascot Bermuda",
        "Aspen Bermuda",
        "AXIS Specialty",
        "Chubb Tempest Reinsurance",
        "Everest Reinsurance Bermuda",
        "Hannover Re Bermuda",
        "Markel Bermuda",
        "Partner Reinsurance Company",
        "Renaissance Reinsurance",
        # Original 20
        "Endurance Specialty Insurance",
        "XL Bermuda",
        "AXA XL Reinsurance",
        "Validus Reinsurance",
        "Somers Re",
        "Lancashire Insurance Company",
        "Hiscox Insurance Company Bermuda",
        "Canopius Reinsurance",
        "Conduit Reinsurance",
        "Fidelis Insurance Bermuda",
        "Fortitude Reinsurance Company",
        "Group Ark Insurance",
        "Hamilton Re",
        "Harrington Re",
        "Liberty Specialty Markets Bermuda",
        "MS Amlin AG",
        "Premia Reinsurance",
        "Starr Insurance & Reinsurance",
        "Vantage Risk",
        "SiriusPoint Bermuda Insurance",
        # Additional 10
        "ABR Reinsurance Ltd.",
        "Allied World Assurance Company Ltd",
        "American International Reinsurance Company Ltd.",
        "Antares Reinsurance Company Limited",
        "Argo Re Ltd.",
        "Brit Reinsurance Bermuda Limited",
        "Convex Re Limited",
        "DaVinci Reinsurance Ltd.",
        "Everest International Reinsurance Ltd.",
        "Fortitude International Reinsurance Ltd."
    ]
}

# Balance Sheet items to extract (with search patterns for flexibility)
BALANCE_SHEET_ITEMS = {
    # Investment breakdown
    "Fixed Maturities - AFS": ["Fixed Maturities", "Available for Sale"],
    "Fixed Maturities - Trading": ["Fixed Maturities", "Trading"],
    "Equity Securities": ["Equity Securities", "Equity"],
    "Short-term Investments": ["Short-term", "Investments"],
    "Other Investments": ["Other", "Investments"],
    # Aggregate
    "Total Investments": ["TOTAL INVESTMENTS"],
    # Other assets
    "Cash and Cash Equivalents": ["Cash", "equivalents"],
    "Total Assets": ["TOTAL ASSETS"],
    # Liabilities & Equity
    "Loss Reserves": ["Loss Reserves"],
    "Unearned Premiums": ["Unearned Premiums"],
    "Total Liabilities": ["TOTAL LIABILITIES"],
    "Total Equity": ["TOTAL EQUITY"]
}

# Income Statement items to extract
INCOME_STATEMENT_ITEMS = {
    # Premiums
    "Gross Premiums Written": ["GROSS PREMIUMS WRITTEN"],
    "Net Premiums Earned": ["NET PREMIUMS EARNED"],
    # Investment income
    "Net Investment Income": ["Net Investment Income"],
    "Investment Gains/Losses": ["Investment Gains", "Investment Losses"],
    # Results
    "Total Revenues": ["TOTAL REVENUES"],
    "Losses and LAE": ["Losses and LAE", "Losses and Loss Adjustment"],
    "Total Expenses": ["TOTAL EXPENSES"],
    "Net Income": ["NET INCOME"]
}

# Cash Flow items to extract
CASH_FLOW_ITEMS = {
    "Operating Cash Flow": ["Operating", "Activities"],
    "Investing Cash Flow": ["Investing", "Activities"],
    "Financing Cash Flow": ["Financing", "Activities"]
}

# ============================================================================
# EXTRACTION LOGIC
# ============================================================================

class MultiYearExtractor:
    """Flexible extractor for multi-year financial data"""

    def __init__(self, config):
        self.config = config
        self.data = {}
        self.extraction_log = []
        self.errors = []

    def extract_all(self):
        """Extract data for all configured years"""
        print("\n" + "=" * 70)
        print("MULTI-YEAR FINANCIAL DATA EXTRACTION")
        print("=" * 70)
        print(f"\nExtracting {len(self.config['years'])} years for {len(self.config['target_companies'])} companies")

        for year in sorted(self.config['years']):
            print(f"\n[{year}] Extracting data...")
            self.extract_year(year)

        print("\n" + "=" * 70)
        print("EXTRACTION COMPLETE")
        print("=" * 70)

        return self.data

    def extract_year(self, year):
        """Extract data for a single year"""
        workbook_path = self.config['workbook_sources'].get(year)

        if not workbook_path:
            msg = f"No workbook source configured for {year}"
            self.errors.append(msg)
            print(f"  ERROR: {msg}")
            return

        if not Path(workbook_path).exists():
            msg = f"Workbook not found: {workbook_path}"
            self.errors.append(msg)
            print(f"  ERROR: {msg}")
            return

        try:
            wb = openpyxl.load_workbook(workbook_path, data_only=True)
            print(f"  Loaded: {workbook_path}")
            print(f"  Sheets: {', '.join(wb.sheetnames)}")

            year_data = {
                "balance_sheet": {},
                "income_statement": {},
                "cash_flows": {}
            }

            # Extract each sheet
            if "Balance Sheet" in wb.sheetnames:
                print(f"  Extracting Balance Sheet...", end=" ")
                year_data["balance_sheet"] = self._extract_sheet(
                    wb["Balance Sheet"],
                    BALANCE_SHEET_ITEMS,
                    year
                )
                print("OK")

            if "Income Statement" in wb.sheetnames:
                print(f"  Extracting Income Statement...", end=" ")
                year_data["income_statement"] = self._extract_sheet(
                    wb["Income Statement"],
                    INCOME_STATEMENT_ITEMS,
                    year
                )
                print("OK")

            if "Cash Flows" in wb.sheetnames:
                print(f"  Extracting Cash Flows...", end=" ")
                year_data["cash_flows"] = self._extract_sheet(
                    wb["Cash Flows"],
                    CASH_FLOW_ITEMS,
                    year
                )
                print("OK")

            # Calculate ratios
            print(f"  Calculating ratios...", end=" ")
            ratios = self._calculate_ratios(year_data["balance_sheet"], year_data["income_statement"])
            year_data["ratios"] = ratios
            print("OK")

            self.data[str(year)] = year_data
            self.extraction_log.append(f"[OK] {year}: {workbook_path}")

        except Exception as e:
            msg = f"Error extracting {year}: {str(e)}"
            self.errors.append(msg)
            print(f"\n  ERROR: {msg}")

    def _extract_sheet(self, worksheet, item_definitions, year):
        """Extract data from a single sheet"""
        data = {}

        # Discover companies in this sheet
        companies = self._discover_companies(worksheet)

        if not companies:
            print(f"\n    Warning: No companies found in sheet")
            return data

        # Extract each item
        for item_name, search_patterns in item_definitions.items():
            row = self._find_row(worksheet, search_patterns)

            if row:
                data[item_name] = {}
                for company_name, col in companies.items():
                    try:
                        value = worksheet.cell(row=row, column=col).value
                        if isinstance(value, (int, float)):
                            data[item_name][company_name] = float(value)
                        elif value is not None:
                            # Try to convert string numbers
                            try:
                                data[item_name][company_name] = float(str(value).replace(',', ''))
                            except:
                                data[item_name][company_name] = 0
                        else:
                            data[item_name][company_name] = 0
                    except:
                        data[item_name][company_name] = 0

        return data

    def _discover_companies(self, worksheet):
        """Discover company names and columns in the worksheet
        Supports two formats:
        - Format A: Companies in row 2, starting from column C (current 40-company workbooks)
        - Format B: Companies in row 1, starting from column B (BMA_Class4 workbooks)
        """
        companies = {}

        # Try Format A first (row 2, starting from column C)
        for col in range(3, 50):  # Start from column C
            company_name = worksheet.cell(row=2, column=col).value
            if company_name:
                # Match against target companies (fuzzy matching)
                for target in self.config['target_companies']:
                    if target.lower() in str(company_name).lower() or str(company_name).lower() in target.lower():
                        companies[target] = col
                        break

        # If no companies found, try Format B (row 1, starting from column B)
        if not companies:
            for col in range(2, 50):  # Start from column B
                company_name = worksheet.cell(row=1, column=col).value
                if company_name and not any(x in str(company_name).lower() for x in ['line item', 'description', 'period', 'date']):
                    # Match against target companies (fuzzy matching)
                    for target in self.config['target_companies']:
                        if target.lower() in str(company_name).lower() or str(company_name).lower() in target.lower():
                            companies[target] = col
                            break

        return companies

    def _find_row(self, worksheet, search_patterns):
        """Find a row by searching column A for patterns"""
        for row in range(1, 100):
            cell_value = worksheet.cell(row=row, column=1).value
            if cell_value:
                cell_str = str(cell_value).upper()
                # Check if ALL patterns match (AND logic)
                if all(pattern.upper() in cell_str for pattern in search_patterns):
                    return row
        return None

    def _calculate_ratios(self, balance_sheet, income_statement):
        """Calculate key financial ratios"""
        ratios = {
            "Loss Ratio (%)": {},
            "Expense Ratio (%)": {},
            "Combined Ratio (%)": {},
            "ROE (%)": {},
            "ROA (%)": {},
            "Equity Ratio (%)": {},
            "Investment Return (%)": {},
            "Investment Yield (%)": {},
            "Investments to Assets (%)": {}
        }

        # Get all companies
        all_companies = set()
        for item_data in balance_sheet.values():
            all_companies.update(item_data.keys())

        for company in all_companies:
            try:
                # ROE = Net Income / Total Equity
                net_income = income_statement.get("Net Income", {}).get(company, 0)
                total_equity = balance_sheet.get("Total Equity", {}).get(company, 1)
                ratios["ROE (%)"][company] = round((net_income / total_equity) * 100, 1) if total_equity else 0

                # ROA = Net Income / Total Assets
                total_assets = balance_sheet.get("Total Assets", {}).get(company, 1)
                ratios["ROA (%)"][company] = round((net_income / total_assets) * 100, 1) if total_assets else 0

                # Equity Ratio = Total Equity / Total Assets
                ratios["Equity Ratio (%)"][company] = round((total_equity / total_assets) * 100, 1) if total_assets else 0

                # Loss Ratio = Losses & LAE / Net Premiums Earned
                losses = income_statement.get("Losses and LAE", {}).get(company, 0)
                net_premiums = income_statement.get("Net Premiums Earned", {}).get(company, 1)
                ratios["Loss Ratio (%)"][company] = round((losses / net_premiums) * 100, 1) if net_premiums and losses else 0

                # Expense Ratio = Total Expenses / Net Premiums Earned
                total_expenses = income_statement.get("Total Expenses", {}).get(company, 0)
                ratios["Expense Ratio (%)"][company] = round((total_expenses / net_premiums) * 100, 1) if net_premiums else 0

                # Combined Ratio = (Losses + Expenses) / Net Premiums Earned
                if net_premiums:
                    combined = (losses + total_expenses) / net_premiums * 100
                    ratios["Combined Ratio (%)"][company] = round(combined, 1)
                else:
                    ratios["Combined Ratio (%)"][company] = 0

                # Investment Return = (Income + Gains) / Investments
                inv_income = income_statement.get("Net Investment Income", {}).get(company, 0)
                inv_gains = income_statement.get("Investment Gains/Losses", {}).get(company, 0)
                total_investments = balance_sheet.get("Total Investments", {}).get(company, 1)
                if total_investments:
                    ratios["Investment Return (%)"][company] = round((inv_income + inv_gains) / total_investments * 100, 2)
                    ratios["Investment Yield (%)"][company] = round(inv_income / total_investments * 100, 2)
                else:
                    ratios["Investment Return (%)"][company] = 0
                    ratios["Investment Yield (%)"][company] = 0

                # Investments to Assets
                ratios["Investments to Assets (%)"][company] = round((total_investments / total_assets) * 100, 1) if total_assets else 0

            except Exception as e:
                for metric in ratios:
                    ratios[metric][company] = 0

        return ratios


def main():
    """Main extraction pipeline"""
    # Extract data
    extractor = MultiYearExtractor(EXTRACTION_CONFIG)
    dashboard_data = extractor.extract_all()

    # Combine with configuration
    output = {
        "companies": EXTRACTION_CONFIG['target_companies'],
        "years": sorted(EXTRACTION_CONFIG['years']),
        "extraction_date": datetime.now().isoformat(),
        "data": dashboard_data
    }

    # Save to JSON
    with open("dashboard_data.json", "w") as f:
        json.dump(output, f, indent=2)

    print(f"\n[OK] dashboard_data.json created successfully!")
    print(f"  - Companies: {len(EXTRACTION_CONFIG['target_companies'])}")
    print(f"  - Years: {', '.join(map(str, sorted(EXTRACTION_CONFIG['years'])))}")
    print(f"  - File size: {Path('dashboard_data.json').stat().st_size / 1024:.1f} KB")

    # Log any errors
    if extractor.errors:
        print(f"\n[WARNING] {len(extractor.errors)} error(s) during extraction:")
        for error in extractor.errors:
            print(f"  - {error}")

    return 0 if not extractor.errors else 1


if __name__ == "__main__":
    exit(main())
