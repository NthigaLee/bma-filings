#!/usr/bin/env python3
"""
Extract 2019-2024 financial data from BMA consolidated workbook.

This script handles the consolidated BMA_Class4_Financials_2020_2024.xlsx
file which contains all companies and years in a single workbook with
consistent formatting across all years.

Format: Columns are "Company Name Year", rows are metrics, values are in cells.
"""

import openpyxl
import json
import re
from pathlib import Path
from typing import Dict, List, Tuple, Optional

# Target companies for extraction (40 companies in current dashboard)
TARGET_COMPANIES = [
    "Arch Reinsurance", "Ascot Bermuda", "Aspen Bermuda", "AXIS Specialty",
    "Chubb Tempest Reinsurance", "Everest Reinsurance Bermuda", "Hannover Re Bermuda",
    "Markel Bermuda", "Partner Reinsurance Company", "Renaissance Reinsurance",
    "Endurance Specialty Insurance", "XL Bermuda", "AXA XL Reinsurance",
    "Validus Reinsurance", "Somers Re", "Lancashire Insurance Company",
    "Hiscox Insurance Company Bermuda", "Canopius Reinsurance", "Conduit Reinsurance",
    "Fidelis Insurance Bermuda", "Fortitude Reinsurance Company", "Group Ark Insurance",
    "Hamilton Re", "Harrington Re", "Liberty Specialty Markets Bermuda", "MS Amlin AG",
    "Premia Reinsurance", "Starr Insurance & Reinsurance", "Vantage Risk",
    "SiriusPoint Bermuda Insurance", "ABR Reinsurance Ltd.", "Allied World Assurance Company Ltd",
    "American International Reinsurance Company Ltd.", "Antares Reinsurance Company Limited",
    "Argo Re Ltd.", "Brit Reinsurance Bermuda Limited", "Convex Re Limited",
    "DaVinci Reinsurance Ltd.", "Everest International Reinsurance Ltd.",
    "Fortitude International Reinsurance Ltd."
]

# Metrics to extract (must match workbook row labels, approximately)
BALANCE_SHEET_ITEMS = {
    "Total Investments": ["Total Investments", "Total investments"],
    "Total Assets": ["Total Assets", "Total assets"],
    "Total Equity": ["Total shareholder", "Total Shareholders Equity", "Total equity"],
    "Total Liabilities": ["Total Liabilities", "Total liabilities"],
    "Cash": ["Cash and cash equivalents", "Cash and cash eq"],
}

INCOME_STATEMENT_ITEMS = {
    "Gross Premiums": ["Gross Premiums", "Gross premiums written"],
    "Net Premiums Earned": ["Net Premiums Earned", "Net premiums earned"],
    "Losses": ["Net losses and loss adjustment", "Losses and loss"],
    "Total Expenses": ["Total Expenses", "Total expenses"],
    "Net Income": ["Net Income", "Net income"],
    "Investment Income": ["Net Investment Income", "Net investment income"],
}


class ConsolidatedExtractor:
    """Extract financial data from consolidated BMA workbook."""

    def __init__(self, workbook_path: str, target_years: List[int] = None):
        self.workbook_path = workbook_path
        self.target_years = target_years or [2021, 2022, 2023, 2024]
        self.data = {}
        self.companies_by_year = {}
        self.extraction_log = []
        self.errors = []

    def extract_all(self) -> Dict:
        """Extract all data from consolidated workbook."""
        print("\n" + "=" * 70)
        print("CONSOLIDATED FINANCIAL DATA EXTRACTION")
        print("=" * 70)

        print(f"\nConfiguration:")
        print(f"  Source: {self.workbook_path}")
        print(f"  Years: {self.target_years}")
        print(f"  Target companies: {len(TARGET_COMPANIES)}")

        try:
            wb = openpyxl.load_workbook(self.workbook_path, data_only=True)
            print(f"\nWorkbook loaded. Sheets: {wb.sheetnames}")

            # Extract from each sheet
            self._extract_sheet(wb, "Balance Sheet", "balance_sheet", BALANCE_SHEET_ITEMS)
            self._extract_sheet(wb, "Income Statement", "income_statement", INCOME_STATEMENT_ITEMS)

            # Calculate ratios
            self._calculate_ratios()

            # Prepare output format
            output = self._prepare_output()

            print("\n" + "=" * 70)
            print("EXTRACTION COMPLETE")
            print("=" * 70)

            return output

        except Exception as e:
            print(f"\nERROR: {e}")
            import traceback
            traceback.print_exc()
            raise

    def _extract_sheet(self, wb, sheet_name: str, data_type: str, metric_map: Dict):
        """Extract a specific sheet from the workbook."""
        print(f"\n[{sheet_name}] Extracting...")

        if sheet_name not in wb.sheetnames:
            print(f"  ERROR: Sheet '{sheet_name}' not found")
            return

        ws = wb[sheet_name]

        # Parse column headers to find company-year combinations
        headers = []
        for col in range(2, 100):  # Start from column B
            cell_value = ws.cell(1, col).value
            if not cell_value:
                break
            headers.append((col, cell_value))

        print(f"  Found {len(headers)} company-year columns")

        # Parse each metric row
        for row in range(2, ws.max_row + 1):
            metric_label = ws.cell(row, 1).value

            if not metric_label:
                continue

            # Find matching metric in our target list
            metric_name = self._match_metric(metric_label, metric_map)

            if not metric_name:
                continue

            # Initialize metric data structure
            if metric_name not in self.data:
                self.data[metric_name] = {}

            # Extract values for each company-year
            for col, header_value in headers:
                company_name, year = self._parse_header(header_value)

                if not company_name or year not in self.target_years:
                    continue

                if year not in self.data[metric_name]:
                    self.data[metric_name][year] = {}

                # Get value
                value = ws.cell(row, col).value
                if value is None:
                    value = 0

                # Convert to float if numeric
                try:
                    if isinstance(value, str):
                        # Remove currency symbols and commas
                        value = float(value.replace('$', '').replace(',', '').strip())
                    else:
                        value = float(value) if value else 0
                except (ValueError, AttributeError):
                    value = 0

                self.data[metric_name][year][company_name] = value

    def _parse_header(self, header: str) -> Tuple[Optional[str], Optional[int]]:
        """Parse 'Company Name Year' format from header."""
        if not header:
            return None, None

        header = str(header).strip()

        # Try to extract year (look for 4-digit number at end)
        match = re.search(r'(\d{4})$', header)
        if not match:
            return None, None

        year = int(match.group(1))
        company = header[:match.start()].strip()

        # Normalize company name
        company = self._normalize_company_name(company)

        return company, year

    def _normalize_company_name(self, name: str) -> str:
        """Normalize company names for matching."""
        if not name:
            return name

        # Exact matches
        if name in TARGET_COMPANIES:
            return name

        # Try fuzzy matching
        for target in TARGET_COMPANIES:
            if target.lower() in name.lower() or name.lower() in target.lower():
                return target

        return name

    def _match_metric(self, label: str, metric_map: Dict) -> Optional[str]:
        """Match metric label to our target metrics."""
        if not label:
            return None

        label_lower = str(label).lower()

        for metric_name, patterns in metric_map.items():
            for pattern in patterns:
                if pattern.lower() in label_lower:
                    return metric_name

        return None

    def _calculate_ratios(self):
        """Calculate financial ratios."""
        print(f"\n[Ratios] Calculating...")

        if "ratios" not in self.data:
            self.data["ratios"] = {}

        # For each year and company, calculate ratios
        for year in self.target_years:
            if year not in self.data.get("ratios", {}):
                self.data["ratios"][year] = {}

            # Get all companies for this year
            companies = set()
            for metric in self.data.values():
                if isinstance(metric, dict) and year in metric:
                    companies.update(metric[year].keys())

            # Calculate ratios for each company
            for company in companies:
                # Loss Ratio = Losses / Net Premiums Earned
                losses = self.data.get("Losses", {}).get(year, {}).get(company, 0)
                npe = self.data.get("Net Premiums Earned", {}).get(year, {}).get(company, 1)
                loss_ratio = (losses / npe * 100) if npe > 0 else 0

                # ROE = Net Income / Total Equity
                ni = self.data.get("Net Income", {}).get(year, {}).get(company, 0)
                eq = self.data.get("Total Equity", {}).get(year, {}).get(company, 1)
                roe = (ni / eq * 100) if eq > 0 else 0

                # ROA = Net Income / Total Assets
                ta = self.data.get("Total Assets", {}).get(year, {}).get(company, 1)
                roa = (ni / ta * 100) if ta > 0 else 0

                # Store ratios
                if company not in self.data["ratios"][year]:
                    self.data["ratios"][year][company] = {}

                self.data["ratios"][year][company]["Loss Ratio %"] = round(loss_ratio, 2)
                self.data["ratios"][year][company]["ROE %"] = round(roe, 2)
                self.data["ratios"][year][company]["ROA %"] = round(roa, 2)

    def _prepare_output(self) -> Dict:
        """Prepare output in dashboard format."""
        # Get all companies across all years
        all_companies = set()
        for metric in self.data.values():
            if isinstance(metric, dict):
                for year_data in metric.values():
                    if isinstance(year_data, dict):
                        all_companies.update(year_data.keys())

        all_companies = sorted(list(all_companies))

        # Organize by year
        output_data = {}
        for year in sorted(self.target_years):
            output_data[str(year)] = {}

            for metric_name, metric_data in self.data.items():
                if year in metric_data:
                    output_data[str(year)][metric_name] = metric_data[year]
                else:
                    output_data[str(year)][metric_name] = {}

        return {
            "companies": all_companies,
            "years": sorted(self.target_years),
            "data": output_data
        }


def main():
    """Main extraction process."""
    extractor = ConsolidatedExtractor(
        'data/BMA_Class4_Financials_2020_2024.xlsx',
        target_years=[2021, 2022, 2023, 2024]
    )

    result = extractor.extract_all()

    # Save to file
    output_file = 'dashboard_data_historical_2021_2024.json'
    with open(output_file, 'w') as f:
        json.dump(result, f, indent=2)

    print(f"\nData saved to: {output_file}")
    print(f"Companies: {len(result['companies'])}")
    print(f"Years: {result['years']}")

    # Show sample data
    if result['data']:
        first_year = str(result['years'][0])
        if first_year in result['data']:
            year_data = result['data'][first_year]
            print(f"\nSample data from {first_year}:")
            for metric in list(year_data.keys())[:3]:
                companies_with_data = [c for c, v in year_data[metric].items() if v > 0]
                print(f"  {metric}: {len(companies_with_data)} companies with data")


if __name__ == '__main__':
    main()
